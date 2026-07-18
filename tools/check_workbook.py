from __future__ import annotations

import argparse
import json
import re
from pathlib import Path

from openpyxl import load_workbook


parser = argparse.ArgumentParser(description="Validate a generated pump-down workbook.")
parser.add_argument("workbook", type=Path, help="Generated .xlsx file")
parser.add_argument("--expected-readings", type=int, help="Optional exact pressure-reading count")
parser.add_argument("--expected-transitions", type=int, help="Optional exact equipment-transition count")
args = parser.parse_args()

path = args.workbook.expanduser().resolve()
if not path.is_file():
    raise FileNotFoundError(path)
workbook = load_workbook(path, data_only=False, read_only=False)
expected = ["Summary", "Pressure Data", "Equipment States", "Events & Notes", "Chart Data"]
if workbook.sheetnames != expected:
    raise AssertionError(f"Unexpected sheets: {workbook.sheetnames}")

pressure = workbook["Pressure Data"]
events = workbook["Events & Notes"]
summary = workbook["Summary"]
headers = [pressure.cell(1, column).value for column in range(1, pressure.max_column + 1)]
required_states = [
    "Pumps Power ON",
    "Turbo Rotor ON",
    "Turbo Vent OPEN",
    "Relay 1 CLOSED",
    "Turbo Gate CLOSED",
    "Turbo Gate OPEN",
    "Argon Gate OPEN",
    "Argon Gate CLOSED",
]
for state in required_states:
    if state not in headers:
        raise AssertionError(f"Missing state column: {state}")
reading_count = pressure.max_row - 1
if reading_count < 1:
    raise AssertionError("Pressure Data does not contain any readings")
if args.expected_readings is not None and reading_count != args.expected_readings:
    raise AssertionError(
        f"Expected {args.expected_readings} pressure readings, found {reading_count}"
    )

transition_count = sum(
    1 for row in events.iter_rows(min_row=2, values_only=True) if row[4] == "Equipment state"
)
if args.expected_transitions is not None and transition_count != args.expected_transitions:
    raise AssertionError(
        f"Expected {args.expected_transitions} equipment transitions, found {transition_count}"
    )
if len(summary._charts) != 1:
    raise AssertionError(f"Expected one summary chart, found {len(summary._charts)}")
chart = summary._charts[0]
if float(chart.y_axis.scaling.logBase) != 10.0:
    raise AssertionError("Summary chart does not use a base-10 logarithmic y-axis")

encoded_state_pattern = re.compile(r"(?<![A-Za-z0-9])[01]{8}(?![A-Za-z0-9])")
for sheet in workbook.worksheets:
    for row in sheet.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and encoded_state_pattern.search(cell.value):
                raise AssertionError(f"Encoded state leaked to {sheet.title}!{cell.coordinate}")

print(
    json.dumps(
        {
            "sheets": workbook.sheetnames,
            "pressure_rows": reading_count,
            "equipment_transitions": transition_count,
            "summary_charts": len(summary._charts),
            "log_base": chart.y_axis.scaling.logBase,
        }
    )
)
workbook.close()
